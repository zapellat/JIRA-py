import os
import re
import json
import logging
import openpyxl
from atlassian import Jira
from requests.exceptions import HTTPError
from datetime import datetime, time
from openpyxl.styles import NamedStyle
from dotenv import load_dotenv

# LOAD ENV #
load_dotenv()

# VARIABLES AND GLOBAL CONFIG #
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

# JIRA API CONFIG #
JIRA_URL = os.getenv('JIRA_URL')
JIRA_USERNAME = os.getenv('JIRA_USERNAME')
JIRA_TOKEN = os.getenv('JIRA_TOKEN')

jira_api = Jira(
    url=JIRA_URL, username=JIRA_USERNAME,
    token=JIRA_TOKEN
)
logging.info("JIRA API Configurada e Autenticada.")

# SEARCHING PROJECTS #
search_terms = ["ProjectName1", "ProjectName2"]
all_projects = jira_api.projects()

matching_projects = [
    p["key"]
    for p in all_projects
    if any(term.lower() in p["name"].lower() for term in search_terms)
]

if not matching_projects:
    logging.warning(f"Zero projects found using {search_terms}.")
else:
    logging.info(f"{len(matching_projects)} projects found.")

# CONSTRUINDO JQL #
type_keys = ['TypeKey1', 'TypeKey2', 'TypeKey3']
quoted_types = [f'"{type_key}"' for type_key in type_keys]
type_keys_str = ', '.join(quoted_types)
jql_query = f'Project IN ({", ".join(matching_projects)}) AND Type IN ({type_keys_str})'

logging.info(f"Running Query: {jql_query}")

start_at_index = 0
max_results_per_query = 150
all_jira_issues = []


# EXTRACT JIRA ISSUES #
def fetch_issues(jql, start_at, max_results):
    while True:
        try:
            issues = jira_api.jql(jql, start=start_at, limit=max_results)
            if not issues['issues']:
                break
            all_jira_issues.extend(issues['issues'])
            start_at += len(issues['issues'])
            logging.info(f"Issue(s) found: {len(issues['issues'])}.")
        except HTTPError as err:
            logging.error(f"Error HTTP: {err.response.status_code} - {err.response.text}")
            break
        except Exception as e:
            logging.error(f"Erros: {str(e)}")
            break


# STATUS MAPPING AND GROUPING NAMES (IN PT-BR) #
def map_de_para_status(status):
    status_mapping = {
        'OPEN': 'NÃO INICIADA',
        'TO DO': 'NÃO INICIADA',
        'REVIEW': 'ANALISE',
        'DRAFT': 'RASCUNHO',
        'DEVELOPMENT COMPLETED': 'IMPLEMENTADA',
        'DONE': 'ENCERRADA',
        'IN ANALYSIS': 'ANALISE',
        'ANALISYS APPROVED': 'ANALISE',
        'ESTIMATE': 'ESTIMATIVA',
        'QUOTED': 'ESTIMATIVA',
        'QUOTED APPROVED': 'ESTIMATIVA',
        'WAITING FOR ESTIMATE': 'ESTIMATIVA',
        'WAITING EVALUATION': 'ESTIMATIVA',
        'QUOTE SENT': 'ESTIMATIVA',
        'ESTIMATION DONE': 'ESTIMATIVA',
        'SELECTED FOR DEVELOPMENT': 'DESENVOLVIMENTO',
        'DEVELOPMENT AUTHORIZED': 'DESENVOLVIMENTO',
        'IN DEVELOPMENT': 'DESENVOLVIMENTO',
        'IN PROGRESS': 'DESENVOLVIMENTO',
        'DEVELOPMENT IN PROGRESS': 'DESENVOLVIMENTO',
        'IN QA': 'DESENVOLVIMENTO',
        'WAITING FOR QA': 'DESENVOLVIMENTO',
        'WAITING FOR UAT': 'UAT',
        'RELEASED NOT IN PRODUCTION': 'UAT',
        'FIRST DELIVERY/RELEASED TO TEST': 'IMPLEMENTADA',
        'VERIFIY OK': 'IMPLEMENTADA',
        'FINAL DELIVERY / RELEASED': 'IMPLEMENTADA',
        'IN UAT': 'EM TESTE',
        'READY FOR PRODUCTION': 'UAT FINALIZADO',
        'UAT DONE': 'UAT FINALIZADO',
        'CLOSED': 'ENCERRADA',
        'PRODUCTION APPROVED': 'IMPLEMENTADA',
        'AWAITING THIRD PARTY': 'DESENVOLVIMENTO',
        'CANCELED': 'CANCELADA',
        'CANCELLED': 'CANCELADA',
        'REJECTED': 'CANCELADA',
        'DEFERRED': 'CANCELADA',
        'BONIFICA': 'ESTIMATIVA',
        'RELEASED': 'ENCERRADA',
        'PENDING': 'OUTRO',
        'CONFIRM RELEASED': 'ENCERRADA',
        'RELEASE COMPLETED': 'ENCERRADA',
        'DEFINITION': 'NÃO INICIADA',
        'WAITING APPROVAL': 'APROVAÇÃO',
        'APPROVED': 'APROVADA',
        'APPROVED BY BUSINESS': 'APROVADA',
        'WAITING FOR CONTRACT APPROVAL': 'APROVAÇÃO',
        'ESTIMATION APPROVAL': 'ESTIMATIVA'
    }
    normalized_status = status.upper().strip()
    return status_mapping.get(normalized_status, 'OUTRO')


# FIELDS #
def process_issues(issues):
    processed_issues_list = []
    for issue in issues:
        fields = issue.get('fields', {})
        components = ', '.join([comp['name'] for comp in fields.get('components', [])])
        project_name = fields.get('project', {}).get('name', 'N/A')
        assignee = fields.get('assignee', {})
        status = fields.get('status', {})
        status_name = status.get('name', 'N/A')
        status_description = map_de_para_status(status_name)
        issue_key = issue.get('key', 'N/A')
        summary = fields.get('summary', 'N/A')

        create_dt = fields.get('created')
        create_dt = create_dt if create_dt else '1999-01-01T00:00:00.000+0000'
        create_dt = datetime.strptime(create_dt, "%Y-%m-%dT%H:%M:%S.%f%z").replace(tzinfo=None)

        update_dt = fields.get('updated')
        update_dt = update_dt if update_dt else '1999-01-01T00:00:00.000+0000'
        update_dt = datetime.strptime(update_dt, "%Y-%m-%dT%H:%M:%S.%f%z").replace(tzinfo=None)

        start_dt = fields.get('customfield_10201')
        start_dt = start_dt if start_dt else '1999-01-01'
        start_dt = datetime.strptime(start_dt, "%Y-%m-%d").replace(tzinfo=None)

        end_dt = fields.get('customfield_10200')
        end_dt = end_dt if end_dt else '1999-01-01'
        end_dt = datetime.strptime(end_dt, "%Y-%m-%d").replace(tzinfo=None)

        actual_dev_dt = fields.get('customfield_15475')
        actual_dev_dt = actual_dev_dt if actual_dev_dt else '1999-01-01'
        actual_dev_dt = datetime.strptime(actual_dev_dt, "%Y-%m-%d").replace(tzinfo=None)

        issue_type = fields.get('issuetype', {}).get('name', 'N/A')
        priority = fields.get('priority', {}).get('name', 'N/A') if isinstance(fields.get('priority', {}),
                                                                                   dict) else 'N/A'
        providers = ', '.join([provider['value'] for provider in fields.get('customfield_15338', [])]) if fields.get(
                'customfield_15338') else 'N/A'
        pais_value = fields.get('customfield_11700')
        pais = pais_value['value'] if pais_value else 'N/A'
        labels = ', '.join(fields.get('labels', []))

        contract_id_field = fields.get('customfield_16427')
        contract_id = contract_id_field if contract_id_field else 'N/A'

        providers_issue_field = fields.get('customfield_21508')
        if providers_issue_field is not None:
            providers_issue_summary = providers_issue_field.get('fields', {}).get('summary', 'N/A')
            providers_issue = providers_issue_summary if providers_issue_summary != 'N/A' else 'N/A'
        else:
            providers_issue = 'N/A'

        providers_issue = providers_issue.replace("SISTEMAS COMERCIALES DE BRASIL TECNOLOGIA DA INFORMAÐ£Â‡Ð£ÂƒO",
                                                      "AYESA")
        providers_issue = providers_issue.replace("ENGINEERING DO BRASIL S A", "ENGINEERING DO BRASIL")
        providers_issue = providers_issue.replace("Engineering", "ENGINEERING DO BRASIL")

        apm_tcr_value = fields.get('customfield_14613')
        apm_tcr = ', '.join(apm_tcr_value) if isinstance(apm_tcr_value, list) else apm_tcr_value
        apm_bcr_value = fields.get('customfield_37200')
        apm_bcr = ', '.join(apm_bcr_value) if isinstance(apm_bcr_value, list) else apm_bcr_value

        if apm_tcr and apm_bcr:
            apm = f"{apm_tcr} - {apm_bcr}"
        elif apm_tcr:
            apm = apm_tcr
        elif apm_bcr:
            apm = apm_bcr
        else:
            apm = "N/A"

        # Extract the custom field "Contract Manager"
        contract_manager_field = fields.get('customfield_34700')
        if contract_manager_field:
            if isinstance(contract_manager_field, list):
                contract_manager = ", ".join([user.get('displayName', 'N/A') for user in contract_manager_field])
            elif isinstance(contract_manager_field, dict):
                contract_manager = contract_manager_field.get('displayName', 'N/A')
            else:
                contract_manager = str(contract_manager_field)
        else:
            contract_manager = 'N/A'

        # Extract the custom field "Contract Delegate"
        contract_delegate_field = fields.get('customfield_37500')
        if contract_delegate_field:
            if isinstance(contract_delegate_field, list):
                contract_delegate = ", ".join([user.get('displayName', 'N/A') for user in contract_delegate_field])
            elif isinstance(contract_delegate_field, dict):
                contract_delegate = contract_delegate_field.get('displayName', 'N/A')
            else:
                contract_delegate = str(contract_delegate_field)
        else:
            contract_delegate = 'N/A'

        # Extract the custom field "Approval Need"
        approval_need_field = fields.get('customfield_30001')
        if approval_need_field:
            if isinstance(approval_need_field, dict):
                approval_need = approval_need_field.get('value', 'N/A')
            else:
                approval_need = str(approval_need_field)
        else:
            approval_need = 'N/A'

        assignee_name = assignee['displayName'] if assignee else 'Unassigned'

        rfc_source_environment = issue['fields'].get('customfield_24703', 'NO STATUS')
        rfc_target_environment = issue['fields'].get('customfield_24704', 'NO STATUS')

        rfc_status = issue['fields'].get('customfield_25919', 'NO STATUS')
        rfc_change_type = issue['fields'].get('customfield_16423', 'NA')
        rfc_release_type = (issue['fields'].get('customfield_22203') or {}).get('value', '')
        rfc_apm_name = issue['fields'].get('customfield_16424', 'NA')
        rfc_provider_name = issue['fields'].get('customfield_16426', 'NA')
        rfc_contract_id = issue['fields'].get('customfield_16427', 'NA')
        rfc_auth_note = issue['fields'].get('customfield_16406', 'NA')

        rfc_contents_key_field = fields.get('customfield_38400', [])
        if rfc_contents_key_field:
            rfc_contents_key = rfc_contents_key_field[0].get('key', 'N/A')
            rfc_contents_key_value = rfc_contents_key if rfc_contents_key != 'N/A' else 'N/A'
        else:
            rfc_contents_key_value = 'N/A'

        rfc_contents_field_type = fields.get('customfield_38400', [])
        if rfc_contents_field_type:
            rfc_contents_type = rfc_contents_field_type[0].get('fields', {}).get('issuetype', {}).get('name', 'N/A')
            rfc_contents_type_value = rfc_contents_type if rfc_contents_type != 'N/A' else 'N/A'
        else:
            rfc_contents_type_value = 'N/A'

        rfc_approvedby_field = issue['fields'].get('customfield_12015')
        rfc_approvedby = rfc_approvedby_field['displayName'] if rfc_approvedby_field else "N/A"

        rfc_target_start = (datetime.strptime(fields.get('customfield_10402') or '1999-01-01', '%Y-%m-%d')).strftime(
            '%d/%m/%Y')
        rfc_target_end = (datetime.strptime(fields.get('customfield_10403') or '1999-01-01', '%Y-%m-%d')).strftime(
            '%d/%m/%Y')

        rfc_target_start_time = (datetime.strptime(fields.get('customfield_22102') or '00:00', "%H:%M")).strftime(
            '%H:%M')
        rfc_target_end_time = (datetime.strptime(fields.get('customfield_22101') or '00:00', "%H:%M")).strftime(
            '%H:%M')

        rfc_description = issue['fields']['description']
        rfc_reporter_name= issue['fields']['reporter']['displayName']

        rfc_rollback_field = issue['fields'].get('customfield_21600')
        rfc_rollback = rfc_rollback_field['value'] if rfc_rollback_field else "N/A"

        processed_issues_list.append({
            'issue_key': issue_key,
            'issue_type': issue_type,
            'project_name': project_name,
            'summary': summary,
            'priority': priority,
            'status': status_name.upper(),
            'components': components,
            'providers_issue': providers_issue,
            'create_dt': create_dt,
            'update_dt': update_dt,
            'status_description': status_description.upper(),
            'assignee': assignee_name,
            'providers': providers.upper(),
            'pais': pais,
            'labels': labels,
            'apm': apm,
            'contract_manager': contract_manager,
            'contract_delegate': contract_delegate,
            'approval_need': approval_need,
            'rfc_target_start': rfc_target_start,
            'rfc_target_end': rfc_target_end,
            'rfc_target_start_time':rfc_target_start_time,
            'rfc_target_end_time': rfc_target_end_time,
            'rfc_source_environment': rfc_source_environment,
            'rfc_target_environment': rfc_target_environment,
            'rfc_status': rfc_status,
            'rfc_change_type': rfc_change_type,
            'rfc_release_type': rfc_release_type,
            'rfc_apm_name': rfc_apm_name,
            'rfc_provider_name': rfc_provider_name,
            'rfc_contract_id': rfc_contract_id,
            'rfc_auth_note': rfc_auth_note,
            'rfc_contents_key_value': rfc_contents_key_value,
            'rfc_contents_type_value': rfc_contents_type_value,
            'rfc_approvedby': rfc_approvedby,
            'rfc_description': rfc_description,
            'rfc_reporter_name': rfc_reporter_name,
            'rfc_rollback': rfc_rollback,
            'contract_id': contract_id
        })
    logging.info("Campos processados com sucesso.")
    return processed_issues_list


# SAVING EXCEL FILE #
def save_to_excel(issues, filename='Issue_Extraction.xlsx'):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    issues = sorted(issues, key=lambda x: x['issue_key'])

    # Cria as abas
    ws_jiras = wb.create_sheet(title="ISSUESTAB1")
    ws_cancelados = wb.create_sheet(title="ISSUESTAB2")
    ws_implementados = wb.create_sheet(title="ISSUESTAB3")
    ws_support_esp = wb.create_sheet(title="ISSUESTAB4")

    headers = ['ID', 'Projeto', 'Tipo', 'Sumário', 'Prioridade', 'Status JIRA', 'Status', 'Provider',
               'Assignee', 'Criado Em', 'Atualizado Em', 'Componentes', 'Labels', 'APM List']
    for ws in (ws_jiras, ws_cancelados, ws_implementados, ws_support_esp):
        ws.append(headers)

    jiras_row_index = 2
    cancelados_row_index = 2
    implementados_row_index = 2
    support_esp_row_index = 2

    for issue in issues:
        if issue['issue_type'] not in ("Provider", "RFC"):
            row = [
                issue['issue_key'],
                issue['project_name'],
                issue['issue_type'],
                issue['summary'],
                issue['priority'],
                issue['status'],
                issue['status_description'],
                issue['providers_issue'],
                issue['assignee'],
                issue['create_dt'],
                issue['update_dt'],
                issue['components'],
                issue['labels'],
                issue['apm']
            ]

            if issue['status'] in ['CANCELED', 'REJECTED', 'CANCELLED']:
                current_sheet = ws_cancelados
                row_index = cancelados_row_index
                cancelados_row_index += 1
            elif issue['issue_type'] == "TYPE1":
                current_sheet = ws_support_esp
                row_index = support_esp_row_index
                support_esp_row_index += 1
            elif issue['status'] in ['CONFIRM RELEASED', 'RELEASED', 'RELEASE COMPLETED']:
                current_sheet = ws_implementados
                row_index = implementados_row_index
                implementados_row_index += 1
            else:
                current_sheet = ws_jiras
                row_index = jiras_row_index
                jiras_row_index += 1

            for col_index, item in enumerate(row, start=1):
                cell = current_sheet.cell(row=row_index, column=col_index)
                if isinstance(item, datetime):
                    cell.value = item
                    cell.style = date_style
                elif isinstance(item, time):
                    cell.value = item.strftime('%H:%M')
                else:
                    cell.value = item

    # CREATE TAB 5 - USE IF THE NEW TAB HAVE DIFFERENT HEADER #
    ws_providers = wb.create_sheet(title="ISSUETAB5")
    headers_providers = ['ID', 'Projeto', 'Tipo', 'Sumário', 'Status JIRA', 'Contract ID', 'Assignee', 'Contract Approval Need',
                         'Contract Manager', 'Contract Manager Delegate']
    ws_providers.append(headers_providers)
    providers_row_index = 2

    for issue in issues:
        if issue['issue_type'] == "ISSUETYPE3":
            ws_providers.append([
                issue['issue_key'], issue['issue_type'], issue['project_name'], issue['summary'], issue['status'],
                issue['contract_id'], issue['assignee'], issue['approval_need'], issue['contract_manager'], issue['contract_delegate']
            ])

            if issue['status'] in ['ISSUETAB5']:
                current_sheet = ws_providers
                row_index = providers_row_index
                providers_row_index += 1

    # CREATE TAB 6 - USE IF THE NEW TAB HAVE DIFFERENT HEADER #
    ws_rfc = wb.create_sheet(title="ISSUETASB6")
    headers_rfc = ['ID', 'Projeto', 'Tipo', 'Status RFC', 'Sumário', 'Start', 'Start Time', 'End', 'End Time',
                  'Source', 'Target', 'Rollback', 'Change Type', 'Release Type', 'APM Name', 'Provider', 'Contract', 'Nota CAB', 'Aprovado Por',
                   'Issue', 'Issue Type', 'RFC Descrição', 'Aberto Por']
    ws_rfc.append(headers_rfc)
    rfc_row_index = 2

    for issue in issues:
        if issue['issue_type'] == "ISSUETYPE2":
            ws_rfc.append([
                issue['issue_key'], issue['project_name'], issue['issue_type'], issue['status'], issue['summary'],
                issue['rfc_target_start'], issue['rfc_target_start_time'], issue['rfc_target_end'],
                issue['rfc_target_end_time'],
                ", ".join(issue['rfc_source_environment']) if isinstance(issue['rfc_source_environment'], list) else
                issue['rfc_source_environment'],
                ", ".join(issue['rfc_target_environment']) if isinstance(issue['rfc_target_environment'], list) else
                issue['rfc_target_environment'], issue['rfc_rollback'],
                issue['rfc_change_type'], issue['rfc_release_type'], issue['rfc_apm_name'], issue['rfc_provider_name'],
                issue['rfc_contract_id'],
                issue['rfc_auth_note'], issue['rfc_approvedby'], issue['rfc_contents_key_value'],
                issue['rfc_contents_type_value'],
                issue['rfc_description'], issue['rfc_reporter_name']
            ])

            if issue['status'] in ['ISSUETASB6']:
                current_sheet = ws_rfc
                row_index = rfc_row_index
                rfc_row_index += 1

            for col_index, item in enumerate(row, start=1):
                cell = current_sheet.cell(row=row_index, column=col_index)
                if isinstance(item, datetime):
                    cell.value = item
                    cell.style = date_style
                elif isinstance(item, time):
                    cell.value = item.strftime('%H:%M')
                else:
                    cell.value = item

    wb.save(filename)
    logging.info(f"Issues exported to: '{filename}'")
    os.startfile(filename)


# RUN SCRIPT #
if __name__ == "__main__":
    fetch_issues(jql_query, start_at_index, max_results_per_query)
    processed_issues = process_issues(all_jira_issues)
    save_to_excel(processed_issues)
    logging.info(f"Extracted {len(all_jira_issues)} issues successfully.")
